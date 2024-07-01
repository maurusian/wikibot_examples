from openpyxl import Workbook, load_workbook
from fuzzywuzzy import process

MAX_UNMATCH_SCORE = 74

def extract_rows_from_excel(filename):
    wb = load_workbook(filename)
    ws = wb.active
    data = {}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=2, values_only=True):
        if row[0] is not None and row[1] is not None:
            data[row[0].strip()] = row[1].strip()

    return data

def generate_dict_from_excel(filename):
    wb = load_workbook(filename)
    ws = wb.active
    result_dict = {}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=4, values_only=True):
        qid, village, fraction = row
        fraction = fraction.strip()
        if fraction not in result_dict:
            result_dict[fraction] = {}

        result_dict[fraction][village] = qid

    return result_dict


def get_2004_stats(filename):
    wb = load_workbook(filename)
    ws = wb.active
    result_dict = {}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4, values_only=True):
        population_2004, families_2004, village, fraction = row
        fraction = fraction.strip()
        if fraction not in result_dict:
            result_dict[fraction] = {}

        result_dict[fraction][village] = {"population_2004":population_2004,"families_2004":families_2004}

    return result_dict

def get_closest_match(v, villages):
    closest_match, score = process.extractOne(v, villages)

    print(closest_match, score)

    return closest_match

def write_to_excel(data_list, filename):
    wb = Workbook()
    ws = wb.active

    if not data_list:
        return

    headers = list(data_list[0].keys())
    ws.append(headers)

    for row_dict in data_list:
        row = [row_dict[header] for header in headers]
        ws.append(row)

    wb.save(filename)


#get all fractions that have equivalence
fractions_file = "fractions.xlsx"

#fractions_wb = Workbook(fractions_file)

#fractions_sheet = fractions_wb.active

fractions = extract_rows_from_excel(fractions_file)

#get all Wikidata Qids organized as a dict

village_qid_file = "Villages with qids.xlsx"

wkdt_dict = generate_dict_from_excel(village_qid_file)

#get 2004 stats
stats2004 = get_2004_stats("Al Haouz 2004.xlsx")

#main part

wb = load_workbook("Batch 1.xlsx")
ws = wb.active

new_rows = []

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=7, values_only=True):
    village = row[6]
    fraction = row[5]

    if fraction in fractions.keys():
        fraction_2004 = fractions[fraction]

        potential_list = wkdt_dict[fraction_2004]

        closest_match, score = process.extractOne(village, list(potential_list.keys()))

        if score > MAX_UNMATCH_SCORE:

            qid = wkdt_dict[fraction_2004][closest_match]

        else:
            qid = ""

        if fraction_2004 in stats2004.keys():
            if closest_match in stats2004[fraction_2004].keys():
                population_2004 = stats2004[fraction_2004][closest_match]["population_2004"]
                families_2004 = stats2004[fraction_2004][closest_match]["families_2004"]
            else:
                closest_match_2004, score_2004 = process.extractOne(village, list(stats2004[fraction_2004].keys()))
                if score_2004 > MAX_UNMATCH_SCORE:
                    population_2004 = stats2004[fraction_2004][closest_match_2004]["population_2004"]
                    families_2004 = stats2004[fraction_2004][closest_match_2004]["families_2004"]
                else:
                    population_2004 = ""
                    families_2004 = ""
        else:
            closest_fraction_match_2004, score_fraction_2004 = process.extractOne(fraction_2004, list(stats2004.keys()))

            if score_fraction_2004 > MAX_UNMATCH_SCORE+10:
                closest_match_2004, score_2004 = process.extractOne(closest_match, list(stats2004[closest_fraction_match_2004].keys()))
                if score_2004 > MAX_UNMATCH_SCORE:
                    population_2004 = stats2004[closest_fraction_match_2004][closest_match_2004]["population_2004"]
                    families_2004 = stats2004[closest_fraction_match_2004][closest_match_2004]["families_2004"]
                else:
                    population_2004 = ""
                    families_2004 = ""
            else:
                population_2004 = ""
                families_2004 = ""
        
        #new_rows.append({"village":village,"fraction":fraction,"fraction_2004":fraction_2004,"closest_match":closest_match,"score":score,"qid":qid,"families_2004":families_2004,"population_2004":population_2004})
            
    else:
        fraction_2004 = ""
        closest_match = ""
        score = ""
        qid = ""
        population_2004 = ""
        families_2004 = ""

    new_rows.append({"village":village,"fraction":fraction,"fraction_2004":fraction_2004,"closest_match":closest_match,"score":score,"qid":qid,"families_2004":families_2004,"population_2004":population_2004})
        
write_to_excel(new_rows, "new_file.xlsx")
#print(wkdt_dict)

#print(fractions)

